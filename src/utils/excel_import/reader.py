from typing import Dict, List, Tuple

from openpyxl import Workbook, load_workbook

from .types import ImportErrorDetail

CUSTOMER_SHEET_NAME = "customers"
CUSTOMER_HEADERS = [
    "name",
    "phone",
    "resident_id",
    "address",
    "occupation",
    "driving_type",
    "payment_method",
    "memo",
]

HEADER_ALIASES = {
    "이름": "name",
    "성명": "name",
    "전화번호": "phone",
    "연락처": "phone",
    "주민번호": "resident_id",
    "주민등록번호": "resident_id",
    "직업": "occupation",
    "운전여부": "driving_type",
    "입금방식": "payment_method",
    "메모": "memo",
}


def _normalize_header_name(header: str) -> str:
    text = str(header or "").strip()
    if not text:
        return ""
    lowered = text.lower()
    return HEADER_ALIASES.get(text, HEADER_ALIASES.get(lowered, lowered))


def read_customers_sheet(file_path: str) -> Tuple[List[Dict[str, str]], List[ImportErrorDetail]]:
    errors: List[ImportErrorDetail] = []
    rows: List[Dict[str, str]] = []

    wb = load_workbook(file_path, data_only=True)
    if CUSTOMER_SHEET_NAME not in wb.sheetnames:
        errors.append(
            ImportErrorDetail(
                sheet=CUSTOMER_SHEET_NAME,
                row=1,
                column="sheet",
                error_code="E001",
                message=f"'{CUSTOMER_SHEET_NAME}' sheet is missing",
                action_hint="템플릿 파일을 사용해주세요",
            )
        )
        return rows, errors

    ws = wb[CUSTOMER_SHEET_NAME]
    if ws.max_row < 1:
        errors.append(
            ImportErrorDetail(
                sheet=CUSTOMER_SHEET_NAME,
                row=1,
                column="header",
                error_code="E002",
                message="header row is missing",
                action_hint="템플릿 헤더를 유지해주세요",
            )
        )
        return rows, errors

    raw_headers = [cell.value for cell in ws[1]]
    normalized_headers = [_normalize_header_name(h) for h in raw_headers]
    header_map = {idx: key for idx, key in enumerate(normalized_headers) if key}

    if "name" not in header_map.values() or "phone" not in header_map.values():
        errors.append(
            ImportErrorDetail(
                sheet=CUSTOMER_SHEET_NAME,
                row=1,
                column="header",
                error_code="E003",
                message="required headers(name, phone) are missing",
                action_hint="템플릿 헤더를 수정하지 마세요",
            )
        )
        return rows, errors

    for row_idx in range(2, ws.max_row + 1):
        values = [ws.cell(row=row_idx, column=col_idx + 1).value for col_idx in range(len(raw_headers))]
        if all(v is None or str(v).strip() == "" for v in values):
            continue

        row_data: Dict[str, str] = {"__row_number": row_idx}  # type: ignore[assignment]
        for col_idx, value in enumerate(values):
            key = header_map.get(col_idx)
            if not key:
                continue
            row_data[key] = "" if value is None else str(value).strip()
        rows.append(row_data)

    return rows, errors


def write_customer_template(file_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = CUSTOMER_SHEET_NAME
    ws.append(CUSTOMER_HEADERS)
    ws.append(["홍길동", "010-1234-5678", "900101-1234567", "서울시 강남구", "회사원", "personal", "신용카드", ""])
    wb.save(file_path)

